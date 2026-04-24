"""
Smoke tests ל-storage.save_handler — מוודא שייצוא JSON ו-Excel עובדים
על תוצאה טיפוסית של שרטוט, כולל טיפול בתווים לא חוקיים ל-Excel.
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from storage.save_handler import (
    _sanitize_for_excel,
    save_batch_to_excel,
    save_to_excel,
    save_to_json,
)


@pytest.fixture
def sample_drawing():
    """תוצאה טיפוסית של extract_drawing (מינימום השדות שהייצוא משתמש בהם)."""
    return {
        "part_number": "BBLE12345",
        "drawing_number": "BBLE12345",
        "revision": "A",
        "customer": "RAFAEL",
        "material": "AL 6061-T6",
        "quantity": "1",
        "source_filename": "drawing.pdf",
        "process_summary_hebrew": "ציפוי ניקל אלקטרולס.",
        "notes": "Some notes here.",
        "coating_processes": [
            {
                "type_he": "ניקל אלקטרולס",
                "type": "Electroless Nickel",
                "name": "Electroless nickel plating",
                "thickness": "25um",
                "standard": "ASTM B733",
                "rohs": True,
            }
        ],
        "painting_processes": [],
        "standards": ["ASTM B733", "ISO 9001"],
        "master_matches": [
            {
                "coating": {"type_he": "ניקל אלקטרולס"},
                "kind": "coating",
                "matches": [
                    {
                        "master_id": "M0001",
                        "score": 85,
                        "desc": "Electroless Nickel, High Phos",
                        "standard": "ASTM B733",
                        "thickness": "25um",
                    }
                ],
            }
        ],
        "_validation_warnings": [
            {"severity": "LOW", "type": "INFO", "message": "All good"}
        ],
    }


def test_sanitize_strips_control_chars():
    """תווי בקרה שאסורים ב-Excel מוסרים רקורסיבית."""
    data = {
        "a": "bad\x00\x08\x0b\x0c\x1fgood",
        "b": ["nested\x01", {"c": "deep\x7f"}],
        "d": 42,
    }
    cleaned = _sanitize_for_excel(data)
    assert cleaned["a"] == "badgood"
    assert cleaned["b"] == ["nested", {"c": "deep"}]
    assert cleaned["d"] == 42  # לא מחרוזת — לא משתנה


def test_sanitize_preserves_valid_unicode():
    """טקסט עברי ותווים תקינים לא נפגעים."""
    text = "ציפוי ניקל — 25μm ✓"
    assert _sanitize_for_excel(text) == text


def test_save_to_json_roundtrip(sample_drawing, tmp_path):
    path = save_to_json(sample_drawing, tmp_path / "out.json")
    assert path.exists()
    loaded = json.loads(path.read_text(encoding="utf-8"))
    assert loaded["part_number"] == sample_drawing["part_number"]
    assert "_saved_at" in loaded


def test_save_to_excel_creates_expected_sheets(sample_drawing, tmp_path):
    path = save_to_excel(sample_drawing, tmp_path / "out.xlsx")
    assert path.exists()
    wb = load_workbook(path)
    assert "Summary" in wb.sheetnames
    assert "Coatings" in wb.sheetnames
    assert "Master_Matches" in wb.sheetnames
    assert "Standards" in wb.sheetnames


def test_save_to_excel_handles_illegal_chars(sample_drawing, tmp_path):
    """ודא ש-OCR עם תווי בקרה לא שובר את הייצוא."""
    sample_drawing["material"] = "CRES 15-5PH\x0c ±0.2"
    sample_drawing["notes"] = "Line1\x00Line2"
    # אם הסניטיזציה נשברה — זה יזרוק IllegalCharacterError של openpyxl
    path = save_to_excel(sample_drawing, tmp_path / "out.xlsx")
    assert path.exists()


def test_save_batch_to_excel(sample_drawing, tmp_path):
    drawings = [
        sample_drawing,
        {**sample_drawing, "part_number": "BBLE99999", "source_filename": "b.pdf"},
    ]
    path = save_batch_to_excel(drawings, tmp_path / "batch.xlsx")
    assert path.exists()
    wb = load_workbook(path)
    assert "Summary" in wb.sheetnames
    # שורה לכל שרטוט + header
    summary = wb["Summary"]
    assert summary.max_row == 3


def test_save_batch_excel_empty_list(tmp_path):
    """רשימה ריקה — יוצר workbook תקין עם גיליון Summary ריק."""
    path = save_batch_to_excel([], tmp_path / "empty.xlsx")
    assert path.exists()
    wb = load_workbook(path)
    assert "Summary" in wb.sheetnames
